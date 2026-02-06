from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Sheet1'

# Headers
ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['C1'] = 'City'
ws['D1'] = 'Salary'

# Data
data = [
    ['John Doe', 30, 'New York', 50000],
    ['Jane Smith', 25, 'London', 45000],
    ['Bob Johnson', 35, 'Paris', 60000],
    ['Alice Brown', 28, 'Tokyo', 55000]
]

for row_num, row_data in enumerate(data, start=2):
    for col_num, cell_value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Add another sheet
ws2 = wb.create_sheet('Sheet2')
ws2['A1'] = 'Product'
ws2['B1'] = 'Price'
ws2['C1'] = 'Quantity'

products = [
    ['Apple', 1.50, 100],
    ['Banana', 0.75, 200],
    ['Orange', 2.00, 50]
]

for row_num, row_data in enumerate(products, start=2):
    for col_num, cell_value in enumerate(row_data, start=1):
        ws2.cell(row=row_num, column=col_num, value=cell_value)

wb.save('test.xlsx')
